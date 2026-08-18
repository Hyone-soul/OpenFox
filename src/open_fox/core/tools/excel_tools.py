"""Excel 工具集：read_excel / write_excel / edit_excel / list_sheets。

使用 openpyxl 实现读写改 + 公式支持。
所有路径通过 PathGuard 校验，确保只能在白名单目录内操作。
"""

from __future__ import annotations

from pathlib import Path

from open_fox.core.exceptions import PathGuardViolation
from open_fox.core.security.path_guard import PathGuard
from open_fox.core.tools.base import BaseTool, ToolResult


def _cell_to_value(cell) -> dict:
    """将 openpyxl Cell 转为可序列化的 dict。"""
    result = {"cell": cell.coordinate, "value": cell.value}
    # 公式单元格：value 是以 = 开头的字符串，data_type == 'f'
    if cell.data_type == "f":
        result["formula"] = cell.value
        # 尝试获取计算值（openpyxl 不自动计算公式，仅在加载时缓存）
        # cached 值存在 cell._value 中（openpyxl 内部）
    return result


def _range_to_cells(ws, cell_range: str | None) -> list[str]:
    """将范围字符串转为单元格坐标列表。"""
    if cell_range is None:
        # 全部单元格
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                cells.append(cell.coordinate)
        return cells
    # 指定范围，如 A1:D10
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    cells = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cells.append(cell.coordinate)
    return cells


class ReadExcelTool(BaseTool):
    """读取 Excel 文件内容。"""

    name = "read_excel"
    description = "读取白名单目录下 Excel 文件的单元格数据，支持指定工作表和范围。"
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Excel 文件路径"},
            "sheet": {"type": "string", "description": "工作表名（默认活动表）"},
            "range": {"type": "string",
                      "description": "单元格范围，如 A1:D10（不传则读取全部）"},
        },
        "required": ["path"],
    }

    def __init__(self, path_guard: PathGuard):
        self._guard = path_guard

    def execute(self, **kwargs) -> ToolResult:
        import openpyxl

        try:
            p = self._guard.resolve(kwargs["path"])
        except PathGuardViolation as e:
            return ToolResult(success=False, error=str(e))
        if not p.exists():
            return ToolResult(success=False, error=f"文件不存在：{p}")
        if not p.suffix.lower() in (".xlsx", ".xlsm"):
            return ToolResult(success=False,
                              error=f"仅支持 .xlsx/.xlsm 格式，当前：{p.suffix}")

        try:
            wb = openpyxl.load_workbook(p, data_only=False)
        except Exception as e:
            return ToolResult(success=False, error=f"加载失败：{e}")

        sheet_name = kwargs.get("sheet")
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            return ToolResult(success=False, error="工作簿中没有工作表")

        cell_range = kwargs.get("range")
        cells = _range_to_cells(ws, cell_range)

        # 构建行列表（每行一个 dict）
        rows = []
        current_row = None
        row_data = []
        for coord in cells:
            cell = ws[coord]
            row_num = cell.row
            if current_row is None:
                current_row = row_num
            if row_num != current_row:
                if row_data:
                    rows.append(row_data)
                row_data = []
                current_row = row_num
            row_data.append(_cell_to_value(cell))
        if row_data:
            rows.append(row_data)

        wb.close()

        import json
        result = {
            "sheet": ws.title,
            "rows": len(rows),
            "cols": max(len(r) for r in rows) if rows else 0,
            "data": rows,
        }
        return ToolResult(success=True, content=json.dumps(result, ensure_ascii=False))


class WriteExcelTool(BaseTool):
    """创建或覆盖 Excel 文件。"""

    name = "write_excel"
    description = "在白名单目录下创建或覆盖一个 Excel 文件，支持公式（以 = 开头自动设为公式）。"
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "输出文件路径"},
            "data": {
                "type": "array",
                "description": "二维数组，每个元素是一行（list of cell values）",
                "items": {"type": "array"},
            },
            "sheet": {"type": "string", "description": "工作表名（默认 Sheet1）"},
            "start_cell": {"type": "string",
                           "description": "起始单元格（默认 A1）"},
            "header": {"type": "boolean",
                       "description": "第一行是否为表头（仅影响说明，不改变写入逻辑）"},
        },
        "required": ["path", "data"],
    }

    def __init__(self, path_guard: PathGuard):
        self._guard = path_guard

    def execute(self, **kwargs) -> ToolResult:
        import openpyxl
        from openpyxl.utils import coordinate_to_tuple

        try:
            p = self._guard.resolve(kwargs["path"])
        except PathGuardViolation as e:
            return ToolResult(success=False, error=str(e))

        data = kwargs["data"]
        if not isinstance(data, list) or not data:
            return ToolResult(success=False, error="data 必须是非空二维数组")

        sheet_name = kwargs.get("sheet", "Sheet1")
        start_cell = kwargs.get("start_cell", "A1")

        try:
            if p.exists():
                wb = openpyxl.load_workbook(p)
            else:
                wb = openpyxl.Workbook()
                # 删除默认 Sheet
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

            start_row, start_col = coordinate_to_tuple(start_cell)

            for row_idx, row_data in enumerate(data):
                if not isinstance(row_data, list):
                    row_data = [row_data]
                for col_idx, value in enumerate(row_data):
                    cell = ws.cell(
                        row=start_row + row_idx,
                        column=start_col + col_idx,
                    )
                    # 以 = 开头的字符串自动设为公式
                    if isinstance(value, str) and value.startswith("="):
                        cell.value = value
                    else:
                        cell.value = value

            wb.save(p)
            wb.close()
        except Exception as e:
            return ToolResult(success=False, error=f"写入失败：{e}")

        return ToolResult(
            success=True,
            content=f"已写入：{p}（工作表：{sheet_name}，"
                    f"起始：{start_cell}，行数：{len(data)}）",
        )


class EditExcelTool(BaseTool):
    """批量修改 Excel 单元格。"""

    name = "edit_excel"
    description = "批量修改 Excel 文件中指定单元格的值，支持公式（以 = 开头自动设为公式）。"
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Excel 文件路径"},
            "updates": {
                "type": "array",
                "description": "更新列表，每项含 cell（坐标如 A1）和 value",
                "items": {
                    "type": "object",
                    "properties": {
                        "cell": {"type": "string", "description": "单元格坐标，如 A1"},
                        "value": {"description": "新值（以 = 开头为公式）"},
                    },
                    "required": ["cell", "value"],
                },
            },
            "sheet": {"type": "string", "description": "工作表名（默认活动表）"},
        },
        "required": ["path", "updates"],
    }

    def __init__(self, path_guard: PathGuard):
        self._guard = path_guard

    def execute(self, **kwargs) -> ToolResult:
        import openpyxl

        try:
            p = self._guard.resolve(kwargs["path"])
        except PathGuardViolation as e:
            return ToolResult(success=False, error=str(e))
        if not p.exists():
            return ToolResult(success=False, error=f"文件不存在：{p}")

        updates = kwargs["updates"]
        if not isinstance(updates, list) or not updates:
            return ToolResult(success=False, error="updates 必须是非空列表")

        sheet_name = kwargs.get("sheet")

        try:
            wb = openpyxl.load_workbook(p)
            ws = wb[sheet_name] if sheet_name else wb.active
            if ws is None:
                return ToolResult(success=False, error="工作簿中没有工作表")

            modified = 0
            for upd in updates:
                cell_ref = upd.get("cell", "")
                value = upd.get("value")
                if not cell_ref:
                    continue
                cell = ws[cell_ref]
                if isinstance(value, str) and value.startswith("="):
                    cell.value = value
                else:
                    cell.value = value
                modified += 1

            wb.save(p)
            wb.close()
        except Exception as e:
            return ToolResult(success=False, error=f"修改失败：{e}")

        return ToolResult(
            success=True,
            content=f"已修改 {modified} 个单元格：{p}（工作表：{ws.title}）",
        )


class ListSheetsTool(BaseTool):
    """列出 Excel 文件中的所有工作表。"""

    name = "list_sheets"
    description = "列出白名单目录下 Excel 文件的所有工作表名。"
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Excel 文件路径"},
        },
        "required": ["path"],
    }

    def __init__(self, path_guard: PathGuard):
        self._guard = path_guard

    def execute(self, **kwargs) -> ToolResult:
        import openpyxl

        try:
            p = self._guard.resolve(kwargs["path"])
        except PathGuardViolation as e:
            return ToolResult(success=False, error=str(e))
        if not p.exists():
            return ToolResult(success=False, error=f"文件不存在：{p}")

        try:
            wb = openpyxl.load_workbook(p, read_only=True)
            sheets = wb.sheetnames
            active = wb.active.title if wb.active else None
            wb.close()
        except Exception as e:
            return ToolResult(success=False, error=f"加载失败：{e}")

        import json
        result = {"sheets": sheets, "active": active}
        return ToolResult(success=True, content=json.dumps(result, ensure_ascii=False))
